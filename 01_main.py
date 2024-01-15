import os
import time
from web_functions import *
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup as bs
import requests
import numpy as np
import openpyxl as op
from datetime import datetime


#get no url informado para o webscrapping
req = requests.get("https://gruposeb.gupy.io/")
site = bs(req.content, 'html.parser')
dados = site.findAll('a', attrs={'data-testid': 'job-list__listitem-href'})

#criando um array vazio com a estrutura adequada para receber os dados
arr_vagas = np.empty((len(dados),), dtype=object)

#percorrendo os elementos encontrados
for i, vaga in enumerate(dados):
    #pegar as 'divs' dentro de cada <li>
    divs = vaga.find('div')

    #pegar o texto de cada 'div' e adicionar na lista de vagas
    arr_vagas[i] = [div.get_text(strip=True) for div in divs]

#print no console os dados capturados
for vaga in arr_vagas:
    print(vaga)

#formatar um excel com o retorno do webscrapping
excel_dir = "excel_vagas_seb"
os.makedirs(excel_dir, exist_ok=True)

workbook = op.Workbook()
sheet = workbook.active

head = ["Cargo", "Localidade", "Efetividade"]
sheet.append(head)

for vagas in arr_vagas:
    sheet.append(vagas)

workbook.save(f"{excel_dir}/vagas_seb.xlsx")

time.sleep(5)
#ler excel e preencher microsoft form
#lendo excel
excel = op.load_workbook(f"{excel_dir}/vagas_seb.xlsx")


# Iterar pelas linhas na folha
for row in sheet.iter_rows(min_row=2, values_only=True):
    cargo, localidade, efetividade = row

    # incializar o chrome
    driver_path = ChromeDriverManager().install()
    driver = webdriver.Chrome()

    # acessar site
    url = 'https://forms.office.com/pages/responsepage.aspx?id=QhQEvrbz4UuFCeypyBdSj7tyC7WzU59DoUmUzzgLXidUNllOQ0JYNjVMSDZSN1Q4MUFZWlVXQkw0UC4u'
    driver.get(url)
    time.sleep(2)

    # marcando opção efetivo "sim"
    if efetividade.lower() == "efetivo":
        click_element(driver, css_selector='[data-automation-value="Sim"]')
    else:
        click_element(driver, css_selector='[data-automation-value="Não"]')

    # preenchendo cidade
    xpath_cidade = "//span[contains(text(), 'Cidade')]/ancestor::div[@data-automation-id='questionItem']//input[@data-automation-id='textInput']"
    inject_input(driver, keys=localidade, xpath=xpath_cidade)

    # preenchendo Cargo
    xpath_cargo = "//span[contains(text(), 'Cargo')]/ancestor::div[@data-automation-id='questionItem']//input[@data-automation-id='textInput']"
    inject_input(driver, keys=cargo, xpath=xpath_cargo)

    # tirando print do preenchimento e salvando na pasta 'screenshots_forms'
    screenshot_dir = "screenshots_forms"
    os.makedirs(screenshot_dir, exist_ok=True)

    dt_hr_atual = datetime.now()
    dt_hora_atual = dt_hr_atual.strftime("%d-%m-%Y_%H-%M-%S")
    screenshot_filename = f"{screenshot_dir}/preenchido_{cargo}-{localidade}_{dt_hora_atual}.png"

    driver.save_screenshot(screenshot_filename)
    print(f"Screenshot salvo como: {screenshot_filename}")

    # clicando em enviar
    click_element(driver, css_selector='[data-automation-id="submitButton"]')

    driver.quit()

# Fechar o arquivo Excel
workbook.close()